import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from db import get_connection, row_to_dict

try:
    from crawler_service import normalize_batch_name
except ImportError:
    def normalize_batch_name(name: str) -> str:
        raw = (name or '').strip()
        mapping = {
            '本科一批': '本科批',
            '本科二批': '本科批',
            '本科': '本科批',
            '专科': '专科批',
        }
        return mapping.get(raw, raw or '本科批')

REQUIRED_HEADERS = ['年份', '省份', '批次', '院校代码', '院校名称', '专业代码', '专业名称']

HEADER_MAP = {
    '年份': 'year',
    '省份': 'province',
    '批次': 'batch',
    '科类': 'stream_type',
    '批次备注': 'batch_note',
    '院校代码': 'school_code',
    '院校名称': 'school_name',
    '院校专业组代码': 'school_major_group_code',
    '专业组代码': 'major_group_code',
    '专业组名称': 'major_group_name',
    '院校所在省': 'school_province',
    '城市': 'city',
    '院校类型': 'school_type',
    '办学层次': 'education_level',
    '是否985': 'is_985',
    '是否211': 'is_211',
    '是否双一流': 'is_double_first_class',
    '是否公办': 'is_public',
    '专业代码': 'major_code',
    '专业全称': 'major_name_full',
    '专业名称': 'major_name',
    '专业备注': 'major_remark',
    '专业门类': 'major_category',
    '门类': 'major_category',
    '专业类型': 'major_type',
    '专业类': 'major_type',
    '学历层次': 'degree_type',
    '专业层次': 'degree_type',
    '学制': 'duration',
    '选科要求': 'subject_requirement',
    '招生人数': 'enrollment_count',
    '计划人数': 'enrollment_count',
    '专业组计划人数': 'major_group_plan_count',
    '专业组录取人数': 'major_group_admission_count',
    '是否新增': 'is_new_major',
    '学费': 'tuition',
    '校区': 'campus',
    '特殊说明': 'special_notes',
    '最低分': 'min_score',
    '最低位次': 'min_rank',
    '平均分': 'avg_score',
    '平均位次': 'avg_rank',
    '最高分': 'max_score',
    '最高位次': 'max_rank',
}

INT_FIELDS = {
    'year', 'is_985', 'is_211', 'is_double_first_class', 'is_public',
    'enrollment_count', 'major_group_plan_count', 'major_group_admission_count',
    'tuition', 'min_score', 'min_rank', 'avg_score',
    'avg_rank', 'max_score', 'max_rank',
}

STREAM_TYPES = ('物理', '历史')


def normalize_province_name(value: Any) -> str:
    text = str(value or '').strip()
    for suffix in ('省', '市', '自治区', '壮族自治区', '回族自治区', '维吾尔自治区'):
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def detect_header_row_index(rows: list[tuple[Any, ...] | list[Any]]) -> int:
    for index, row in enumerate(rows[:40]):
        cells = [str(cell).strip() if cell is not None else '' for cell in row]
        if '年份' in cells and '院校代码' in cells and ('专业名称' in cells or '专业全称' in cells):
            return index
    return 0


def validate_headers(headers: list[str]) -> None:
    header_set = {str(header).strip() for header in headers if header}
    if all(header in header_set for header in REQUIRED_HEADERS):
        return

    henan_required = ['年份', '省份', '批次', '院校代码', '院校名称']
    missing = [header for header in henan_required if header not in header_set]
    if missing:
        standard_missing = [header for header in REQUIRED_HEADERS if header not in header_set]
        raise ValueError(f'缺少必填字段：{", ".join(standard_missing)}')
    if '专业名称' not in header_set and '专业全称' not in header_set:
        raise ValueError('缺少必填字段：专业名称 或 专业全称')


def normalize_bool(value: Any, default: int = 0) -> int:
    if value is None or value == '':
        return default
    text = str(value).strip().lower()
    if text in ['1', '是', 'true', 'yes', 'y', '公办', '双一流', '985', '211']:
        return 1
    if text in ['0', '否', 'false', 'no', 'n', '民办']:
        return 0
    return default


def to_int(value: Any, default: int | None = None) -> int | None:
    if value is None or value == '':
        return default
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return default


def normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for cn_key, value in raw.items():
        if cn_key is None:
            continue
        key = HEADER_MAP.get(str(cn_key).strip())
        if not key:
            continue
        if isinstance(value, str):
            value = value.strip()
        row[key] = value

    for field in INT_FIELDS:
        if field in row:
            if field in ['is_985', 'is_211', 'is_double_first_class', 'is_public']:
                row[field] = normalize_bool(row[field], default=1 if field == 'is_public' else 0)
            else:
                row[field] = to_int(row[field])

    row.setdefault('school_province', row.get('province'))
    row.setdefault('city', '')
    row.setdefault('school_type', '')
    row.setdefault('education_level', '本科')
    row.setdefault('is_985', 0)
    row.setdefault('is_211', 0)
    row.setdefault('is_double_first_class', 0)
    row.setdefault('is_public', 1)
    row.setdefault('major_category', '')
    row.setdefault('major_type', '')
    row.setdefault('degree_type', '本科')
    row.setdefault('duration', '')
    return row


def build_major_code(row: dict[str, Any]) -> str:
    raw_code = str(row.get('major_code') or '').strip()
    group_code = str(row.get('school_major_group_code') or '').strip()
    school_code = str(row.get('school_code') or '').strip()
    major_group_code = str(row.get('major_group_code') or '').strip()
    if group_code and raw_code:
        return f'{group_code}-{raw_code}'
    if school_code and major_group_code and raw_code:
        return f'{school_code}{major_group_code}{raw_code}'
    if raw_code:
        return raw_code
    major_name = str(row.get('major_name') or row.get('major_name_full') or 'unknown')
    return f'M-{major_name[:12]}'


def enrich_import_row(row: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(row)
    enriched['province'] = normalize_province_name(enriched.get('province'))
    enriched['school_code'] = str(enriched.get('school_code') or '').strip()
    enriched['school_name'] = str(enriched.get('school_name') or '').strip()
    enriched['major_name'] = str(enriched.get('major_name_full') or enriched.get('major_name') or '').strip()
    enriched['major_code'] = build_major_code(enriched)

    duration = enriched.get('duration')
    if duration is not None and str(duration).strip().isdigit():
        enriched['duration'] = f'{int(duration)}年'

    stream_type = str(enriched.get('stream_type') or '').strip()
    requirement = str(enriched.get('subject_requirement') or '').strip()
    if stream_type in STREAM_TYPES:
        if not requirement or requirement in ('不限', '无要求', '无', '-', '—'):
            enriched['subject_requirement'] = stream_type
        elif stream_type not in requirement:
            enriched['subject_requirement'] = f'{stream_type}；{requirement}'

    note_parts: list[str] = []
    for key, label in (
        ('batch_note', '批次备注'),
        ('major_group_name', '专业组'),
        ('school_major_group_code', '院校专业组代码'),
        ('major_remark', '专业备注'),
        ('is_new_major', '是否新增'),
    ):
        value = enriched.get(key)
        if value is None or value == '':
            continue
        text = str(value).strip()
        if not text:
            continue
        if key == 'major_remark':
            note_parts.append(text.strip('()（）'))
        else:
            note_parts.append(f'{label}：{text}')
    if note_parts:
        existing = str(enriched.get('special_notes') or '').strip()
        merged = '；'.join(note_parts)
        enriched['special_notes'] = f'{existing}；{merged}' if existing else merged

    if enriched.get('enrollment_count') is None and enriched.get('major_group_admission_count') is not None:
        enriched['enrollment_count'] = enriched.get('major_group_admission_count')

    return enriched


def row_has_admission_scores(row: dict[str, Any]) -> bool:
    return any(row.get(field) is not None for field in ('min_score', 'min_rank', 'avg_score', 'avg_rank', 'max_score', 'max_rank'))


def parse_sheet_rows(rows: list[tuple[Any, ...] | list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_index = detect_header_row_index(rows)
    headers = [str(cell).strip() if cell is not None else '' for cell in rows[header_index]]
    validate_headers(headers)
    result: list[dict[str, Any]] = []
    for values in rows[header_index + 1:]:
        if not any(value is not None and str(value).strip() != '' for value in values):
            continue
        raw = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
            if headers[index]
        }
        result.append(enrich_import_row(normalize_row(raw)))
    return result


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    return parse_sheet_rows(rows)


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode('utf-8-sig')
    lines = text.splitlines()
    if not lines:
        return []
    reader_rows = list(csv.reader(io.StringIO(text)))
    if not reader_rows:
        return []
    header_index = detect_header_row_index([tuple(row) for row in reader_rows])
    headers = [cell.strip() for cell in reader_rows[header_index]]
    validate_headers(headers)
    result: list[dict[str, Any]] = []
    for values in reader_rows[header_index + 1:]:
        if not any(str(value).strip() for value in values):
            continue
        raw = {
            headers[index]: values[index] if index < len(values) else ''
            for index in range(len(headers))
            if headers[index]
        }
        result.append(enrich_import_row(normalize_row(raw)))
    return result


def parse_import_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == '.xlsx':
        return parse_xlsx(content)
    if suffix == '.csv':
        return parse_csv(content)
    raise ValueError('仅支持 .xlsx 或 .csv 文件')


def get_or_create_school(connection, row: dict[str, Any]) -> int:
    school = row_to_dict(connection.execute('SELECT school_id FROM schools WHERE school_code = ?', [row['school_code']]).fetchone())
    if school:
        connection.execute(
            '''
            UPDATE schools SET school_name = ?, province = ?, city = ?, school_type = ?, education_level = ?,
              is_985 = ?, is_211 = ?, is_double_first_class = ?, is_public = ?, updated_at = CURRENT_TIMESTAMP
            WHERE school_id = ?
            ''',
            [
                row['school_name'], row.get('school_province'), row.get('city'), row.get('school_type'),
                row.get('education_level'), row.get('is_985'), row.get('is_211'), row.get('is_double_first_class'),
                row.get('is_public'), school['school_id']
            ]
        )
        return school['school_id']

    cursor = connection.execute(
        '''
        INSERT INTO schools (school_code, school_name, province, city, school_type, education_level, is_985, is_211, is_double_first_class, is_public)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        [
            row['school_code'], row['school_name'], row.get('school_province'), row.get('city'), row.get('school_type'),
            row.get('education_level'), row.get('is_985'), row.get('is_211'), row.get('is_double_first_class'), row.get('is_public')
        ]
    )
    return cursor.lastrowid


def get_or_create_major(connection, row: dict[str, Any]) -> int:
    major = row_to_dict(connection.execute('SELECT major_id FROM majors WHERE major_code = ?', [row['major_code']]).fetchone())
    if major:
        connection.execute(
            '''
            UPDATE majors SET major_name = ?, major_category = ?, major_type = ?, degree_type = ?, duration = ?, updated_at = CURRENT_TIMESTAMP
            WHERE major_id = ?
            ''',
            [row['major_name'], row.get('major_category'), row.get('major_type'), row.get('degree_type'), row.get('duration'), major['major_id']]
        )
        return major['major_id']

    cursor = connection.execute(
        '''
        INSERT INTO majors (major_code, major_name, major_category, major_type, degree_type, duration)
        VALUES (?, ?, ?, ?, ?, ?)
        ''',
        [row['major_code'], row['major_name'], row.get('major_category'), row.get('major_type'), row.get('degree_type'), row.get('duration')]
    )
    return cursor.lastrowid


def upsert_plan(connection, row: dict[str, Any], school_id: int, major_id: int) -> None:
    connection.execute(
        '''
        INSERT INTO enrollment_plans (
          year, province, batch, school_id, school_code, major_id, major_code, major_name,
          subject_requirement, enrollment_count, tuition, duration, campus, special_notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(year, province, batch, school_id, major_id) DO UPDATE SET
          subject_requirement = excluded.subject_requirement,
          enrollment_count = excluded.enrollment_count,
          tuition = excluded.tuition,
          duration = excluded.duration,
          campus = excluded.campus,
          special_notes = excluded.special_notes,
          updated_at = CURRENT_TIMESTAMP
        ''',
        [
            row['year'], row['province'], row['batch'], school_id, row['school_code'], major_id, row['major_code'], row['major_name'],
            row.get('subject_requirement'), row.get('enrollment_count'), row.get('tuition'), row.get('duration'), row.get('campus'), row.get('special_notes')
        ]
    )


def upsert_admission(connection, row: dict[str, Any], school_id: int, major_id: int) -> None:
    connection.execute(
        '''
        INSERT INTO admission_records (
          year, province, batch, school_id, school_code, major_id, major_code,
          min_score, min_rank, avg_score, avg_rank, max_score, max_rank, enrollment_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(year, province, batch, school_id, major_id) DO UPDATE SET
          min_score = excluded.min_score,
          min_rank = excluded.min_rank,
          avg_score = excluded.avg_score,
          avg_rank = excluded.avg_rank,
          max_score = excluded.max_score,
          max_rank = excluded.max_rank,
          enrollment_count = excluded.enrollment_count,
          updated_at = CURRENT_TIMESTAMP
        ''',
        [
            row['year'], row['province'], row['batch'], school_id, row['school_code'], major_id, row['major_code'],
            row.get('min_score'), row.get('min_rank'), row.get('avg_score'), row.get('avg_rank'), row.get('max_score'), row.get('max_rank'), row.get('enrollment_count')
        ]
    )


def insert_import_log(connection, import_type: str, file_name: str, total_count: int, success_count: int, fail_count: int, error_message: str | None) -> int:
    cursor = connection.execute(
        '''
        INSERT INTO import_logs (import_type, file_name, total_count, success_count, fail_count, error_message)
        VALUES (?, ?, ?, ?, ?, ?)
        ''',
        [import_type, file_name, total_count, success_count, fail_count, error_message]
    )
    return cursor.lastrowid


def import_admission_rows(filename: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    success_count = 0
    plan_only_count = 0
    errors = []

    with get_connection() as connection:
        for index, row in enumerate(rows, start=2):
            try:
                for field in ['year', 'province', 'batch', 'school_code', 'school_name', 'major_code', 'major_name']:
                    if not row.get(field):
                        raise ValueError(f'第 {index} 行缺少字段：{field}')
                row['batch'] = normalize_batch_name(str(row.get('batch') or ''))
                school_id = get_or_create_school(connection, row)
                major_id = get_or_create_major(connection, row)
                upsert_plan(connection, row, school_id, major_id)
                if row_has_admission_scores(row):
                    upsert_admission(connection, row, school_id, major_id)
                else:
                    plan_only_count += 1
                success_count += 1
            except Exception as exc:
                errors.append(f'第 {index} 行：{exc}')

        fail_count = len(errors)
        error_message = '\n'.join(errors[:20]) if errors else None
        log_id = insert_import_log(connection, 'admission_records', filename, len(rows), success_count, fail_count, error_message)
        connection.commit()

    return {
        'log_id': log_id,
        'total_count': len(rows),
        'success_count': success_count,
        'fail_count': fail_count,
        'plan_only_count': plan_only_count,
        'errors': errors[:20]
    }
