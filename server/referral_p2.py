"""达人分销 P2：物料库、数据看板、海报模板、批量导入、标签、专属福利。"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta
from typing import Any

from db import get_connection, row_to_dict, rows_to_dicts
from referral_p1 import ensure_referral_p1_tables
from referral_service import register_agent, _generate_invite_code, get_default_commission_rate

DEFAULT_MATERIALS = [
    {'category': '话术', 'title': '家长群开场', 'content': '各位家长好，我是XX老师。高考志愿填报关键在「分数+位次+选科」三者结合。我用的智愿填报小程序，能按孩子情况生成冲稳保方案，扫码可领取专属体验。'},
    {'category': '口播', 'title': '30秒口播脚本', 'content': '孩子分数出来了别慌，先看清位次和批次。打开智愿填报，输入分数选科，3分钟出志愿结构，还能看院校专业风险。扫码领取你的专属方案。'},
    {'category': '短视频', 'title': '短视频文案', 'content': '同一分数，有人滑档有人捡漏？差别在志愿结构。智愿填报：测评+报告+智能推荐一条龙。评论区扣「志愿」发你专属入口。'},
]

POSTER_TEMPLATES = [
    {'template_key': 'blue', 'template_name': '经典蓝', 'bg_color': '#1677ff', 'text_color': '#ffffff'},
    {'template_key': 'gold', 'template_name': '金榜题名', 'bg_color': '#d48806', 'text_color': '#ffffff'},
    {'template_key': 'green', 'template_name': '清新绿', 'bg_color': '#389e0d', 'text_color': '#ffffff'},
]

AGENT_TAG_OPTIONS = ['头部', '腰部', '素人', '优质', '待观察']


def _ensure_column(connection, table: str, column: str, ddl: str) -> None:
    cols = [row['name'] for row in connection.execute(f'PRAGMA table_info({table})').fetchall()]
    if column not in cols:
        connection.execute(f'ALTER TABLE {table} ADD COLUMN {column} {ddl}')


def ensure_referral_p2_tables() -> None:
    ensure_referral_p1_tables()
    with get_connection() as connection:
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS referral_materials (
              material_id INTEGER PRIMARY KEY AUTOINCREMENT,
              category TEXT NOT NULL DEFAULT '话术',
              title TEXT NOT NULL,
              content TEXT NOT NULL,
              sort_order INTEGER NOT NULL DEFAULT 0,
              is_active INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            '''
        )
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS referral_poster_templates (
              template_id INTEGER PRIMARY KEY AUTOINCREMENT,
              template_key TEXT NOT NULL UNIQUE,
              template_name TEXT NOT NULL,
              bg_color TEXT NOT NULL DEFAULT '#1677ff',
              text_color TEXT NOT NULL DEFAULT '#ffffff',
              is_active INTEGER NOT NULL DEFAULT 1,
              sort_order INTEGER NOT NULL DEFAULT 0
            )
            '''
        )
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS referral_bonus_claims (
              claim_id INTEGER PRIMARY KEY AUTOINCREMENT,
              user_id INTEGER NOT NULL UNIQUE,
              agent_id INTEGER,
              bonus_beans INTEGER NOT NULL DEFAULT 0,
              bonus_days INTEGER NOT NULL DEFAULT 0,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            '''
        )
        for key, value in [
            ('referral_bonus_beans', '200'),
            ('referral_bonus_days', '3'),
        ]:
            connection.execute(
                '''
                INSERT INTO app_settings (setting_key, setting_value) VALUES (?, ?)
                ON CONFLICT(setting_key) DO NOTHING
                ''',
                [key, value]
            )
        count = connection.execute('SELECT COUNT(*) AS c FROM referral_materials').fetchone()['c']
        if count == 0:
            for index, item in enumerate(DEFAULT_MATERIALS):
                connection.execute(
                    '''
                    INSERT INTO referral_materials (category, title, content, sort_order)
                    VALUES (?, ?, ?, ?)
                    ''',
                    [item['category'], item['title'], item['content'], index + 1]
                )
        _ensure_column(connection, 'referral_poster_templates', 'bg_image_path', "TEXT NOT NULL DEFAULT ''")
        tpl_count = connection.execute('SELECT COUNT(*) AS c FROM referral_poster_templates').fetchone()['c']
        if tpl_count == 0:
            for index, item in enumerate(POSTER_TEMPLATES):
                connection.execute(
                    '''
                    INSERT INTO referral_poster_templates (template_key, template_name, bg_color, text_color, sort_order)
                    VALUES (?, ?, ?, ?, ?)
                    ''',
                    [item['template_key'], item['template_name'], item['bg_color'], item['text_color'], index + 1]
                )
        connection.commit()


def _get_setting(key: str, default: str = '') -> str:
    with get_connection() as connection:
        row = connection.execute('SELECT setting_value FROM app_settings WHERE setting_key = ?', [key]).fetchone()
    return (row['setting_value'] if row else default) or default


def get_bonus_settings() -> dict[str, Any]:
    ensure_referral_p2_tables()
    beans = int(_get_setting('referral_bonus_beans', '200') or 200)
    days = int(_get_setting('referral_bonus_days', '3') or 3)
    return {
        'bonus_beans': beans,
        'bonus_days': days,
        'enabled': beans > 0,
        'reward_text': f'扫码领 {beans} 星鼎豆' if beans > 0 else '扫码领取专属权益',
    }


def get_public_scan_reward() -> dict[str, Any]:
    settings = get_bonus_settings()
    return {
        'bonus_beans': settings['bonus_beans'],
        'bonus_days': settings['bonus_days'],
        'enabled': settings['enabled'],
        'reward_text': settings['reward_text'],
    }


def save_bonus_settings(bonus_beans: int, bonus_days: int) -> dict[str, Any]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        for key, value in [
            ('referral_bonus_beans', str(max(0, int(bonus_beans)))),
            ('referral_bonus_days', str(max(0, int(bonus_days)))),
        ]:
            connection.execute(
                '''
                INSERT INTO app_settings (setting_key, setting_value) VALUES (?, ?)
                ON CONFLICT(setting_key) DO UPDATE SET setting_value = excluded.setting_value, updated_at = CURRENT_TIMESTAMP
                ''',
                [key, value]
            )
        connection.commit()
    return get_bonus_settings()


def list_materials(active_only: bool = True) -> list[dict[str, Any]]:
    ensure_referral_p2_tables()
    sql = 'SELECT * FROM referral_materials WHERE 1=1'
    if active_only:
        sql += ' AND is_active = 1'
    sql += ' ORDER BY sort_order ASC, material_id ASC'
    with get_connection() as connection:
        return rows_to_dicts(connection.execute(sql).fetchall())


def save_material(material_id: int | None, category: str, title: str, content: str, sort_order: int = 0, is_active: int = 1) -> dict[str, Any]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        if material_id:
            connection.execute(
                '''
                UPDATE referral_materials
                SET category = ?, title = ?, content = ?, sort_order = ?, is_active = ?, updated_at = CURRENT_TIMESTAMP
                WHERE material_id = ?
                ''',
                [category, title, content, sort_order, is_active, material_id]
            )
            target_id = material_id
        else:
            cursor = connection.execute(
                '''
                INSERT INTO referral_materials (category, title, content, sort_order, is_active)
                VALUES (?, ?, ?, ?, ?)
                ''',
                [category, title, content, sort_order, is_active]
            )
            target_id = cursor.lastrowid
        connection.commit()
        return row_to_dict(connection.execute('SELECT * FROM referral_materials WHERE material_id = ?', [target_id]).fetchone())


def delete_material(material_id: int) -> None:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        connection.execute('DELETE FROM referral_materials WHERE material_id = ?', [material_id])
        connection.commit()


def list_poster_templates() -> list[dict[str, Any]]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        return rows_to_dicts(connection.execute(
            'SELECT * FROM referral_poster_templates WHERE is_active = 1 ORDER BY sort_order ASC'
        ).fetchall())


def attach_poster_background(template_key: str, bg_filename: str) -> dict[str, Any]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        row = row_to_dict(connection.execute(
            'SELECT * FROM referral_poster_templates WHERE template_key = ?',
            [template_key]
        ).fetchone())
        if not row:
            raise ValueError('海报模板不存在')
        connection.execute(
            '''
            UPDATE referral_poster_templates
            SET bg_image_path = ?
            WHERE template_key = ?
            ''',
            [bg_filename, template_key]
        )
        connection.commit()
        return row_to_dict(connection.execute(
            'SELECT * FROM referral_poster_templates WHERE template_key = ?',
            [template_key]
        ).fetchone())


def save_poster_template(template_key: str, template_name: str, bg_color: str, text_color: str = '#ffffff') -> dict[str, Any]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        connection.execute(
            '''
            INSERT INTO referral_poster_templates (template_key, template_name, bg_color, text_color)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(template_key) DO UPDATE SET
              template_name = excluded.template_name,
              bg_color = excluded.bg_color,
              text_color = excluded.text_color
            ''',
            [template_key, template_name, bg_color, text_color]
        )
        connection.commit()
        return row_to_dict(connection.execute(
            'SELECT * FROM referral_poster_templates WHERE template_key = ?',
            [template_key]
        ).fetchone())


def get_agent_stats(agent_id: int, days: int = 30) -> dict[str, Any]:
    ensure_referral_p2_tables()
    days = max(1, min(int(days or 30), 365))
    since = (datetime.now() - timedelta(days=days)).isoformat(timespec='seconds')
    with get_connection() as connection:
        agent = row_to_dict(connection.execute('SELECT * FROM referral_agents WHERE agent_id = ?', [agent_id]).fetchone())
        if not agent:
            raise ValueError('达人不存在')
        bind_count = connection.execute(
            'SELECT COUNT(*) AS c FROM referral_bindings WHERE agent_id = ? AND bound_at >= ?',
            [agent_id, since]
        ).fetchone()['c']
        scan_logs = connection.execute(
            '''
            SELECT COUNT(*) AS c FROM referral_verify_logs
            WHERE agent_id = ? AND action = 'bind' AND created_at >= ?
            ''',
            [agent_id, since]
        ).fetchone()['c']
        paid_orders = connection.execute(
            '''
            SELECT COUNT(*) AS c, IFNULL(SUM(c.commission_amount), 0) AS amount
            FROM referral_commissions c
            WHERE c.agent_id = ? AND c.status != 'cancelled' AND c.created_at >= ?
            ''',
            [agent_id, since]
        ).fetchone()
        daily = rows_to_dicts(connection.execute(
            '''
            SELECT substr(bound_at, 1, 10) AS day, COUNT(*) AS binds
            FROM referral_bindings
            WHERE agent_id = ? AND bound_at >= ?
            GROUP BY substr(bound_at, 1, 10)
            ORDER BY day ASC
            ''',
            [agent_id, since]
        ).fetchall())
    binds = int(bind_count or 0)
    orders = int(paid_orders['c'] or 0)
    return {
        'days': days,
        'scan_count': int(scan_logs or 0),
        'bind_users': binds,
        'paid_orders': orders,
        'commission_amount': round(float(paid_orders['amount'] or 0), 2),
        'conversion_rate': round((orders / max(binds, 1)) * 100, 1),
        'total_commission': float(agent.get('total_commission') or 0),
        'daily_binds': daily,
    }


def update_agent_profile(agent_id: int, data: dict[str, Any]) -> dict[str, Any]:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        agent = row_to_dict(connection.execute('SELECT * FROM referral_agents WHERE agent_id = ?', [agent_id]).fetchone())
        if not agent:
            raise ValueError('达人不存在')
        connection.execute(
            '''
            UPDATE referral_agents
            SET display_name = COALESCE(?, display_name),
                tags = COALESCE(?, tags),
                douyin_id = COALESCE(?, douyin_id),
                fan_scale = COALESCE(?, fan_scale),
                updated_at = CURRENT_TIMESTAMP
            WHERE agent_id = ?
            ''',
            [
                data.get('display_name'),
                data.get('tags'),
                data.get('douyin_id'),
                data.get('fan_scale'),
                agent_id,
            ]
        )
        connection.commit()
        return row_to_dict(connection.execute('SELECT * FROM referral_agents WHERE agent_id = ?', [agent_id]).fetchone())


def import_agents_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    ensure_referral_p2_tables()
    created = 0
    updated = 0
    errors: list[str] = []
    with get_connection() as connection:
        for index, row in enumerate(rows, start=1):
            phone = str(row.get('phone') or row.get('手机号') or '').strip()
            display_name = str(row.get('display_name') or row.get('昵称') or row.get('博主名') or '').strip()
            douyin_id = str(row.get('douyin_id') or row.get('抖音号') or '').strip()
            fan_scale = str(row.get('fan_scale') or row.get('粉丝量级') or '').strip()
            tags = str(row.get('tags') or row.get('标签') or '').strip()
            if not phone and not display_name:
                errors.append(f'第{index}行：缺少手机号或昵称')
                continue
            user = None
            if phone:
                user = row_to_dict(connection.execute('SELECT * FROM users WHERE phone = ?', [phone]).fetchone())
            if not user:
                cursor = connection.execute(
                    'INSERT INTO users (phone, role, name, openid) VALUES (?, ?, ?, ?)',
                    [phone or None, 'teacher', display_name or f'达人{index}', f'agent_import_{phone or index}_{int(datetime.now().timestamp())}']
                )
                user_id = cursor.lastrowid
            else:
                user_id = user['user_id']
            agent = row_to_dict(connection.execute('SELECT * FROM referral_agents WHERE user_id = ?', [user_id]).fetchone())
            if agent:
                connection.execute(
                    '''
                    UPDATE referral_agents
                    SET display_name = ?, douyin_id = ?, fan_scale = ?, tags = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE agent_id = ?
                    ''',
                    [display_name or agent.get('display_name'), douyin_id, fan_scale, tags, agent['agent_id']]
                )
                updated += 1
            else:
                invite_code = _generate_invite_code(connection)
                connection.execute(
                    '''
                    INSERT INTO referral_agents (user_id, invite_code, display_name, commission_rate, douyin_id, fan_scale, tags)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''',
                    [user_id, invite_code, display_name or f'达人{user_id}', get_default_commission_rate(), douyin_id, fan_scale, tags]
                )
                created += 1
        connection.commit()
    return {'created': created, 'updated': updated, 'errors': errors}


def parse_agents_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(item or '').strip() for item in next(rows_iter, [])]
    result = []
    for row in rows_iter:
        if not row or not any(row):
            continue
        item = {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers)) if headers[i]}
        result.append(item)
    return result


def export_agents_csv() -> str:
    ensure_referral_p2_tables()
    with get_connection() as connection:
        rows = rows_to_dicts(connection.execute(
            '''
            SELECT a.agent_id, a.invite_code, a.display_name, a.tags, a.douyin_id, a.fan_scale,
                   a.commission_rate, a.total_invites, a.total_paid_orders, a.total_commission,
                   u.phone, u.name AS user_name
            FROM referral_agents a
            JOIN users u ON u.user_id = a.user_id
            ORDER BY a.created_at DESC
            '''
        ).fetchall())
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['达人ID', '邀请码', '昵称', '标签', '抖音号', '粉丝量级', '佣金%', '推广人数', '付费单', '累计佣金', '手机号'])
    for item in rows:
        writer.writerow([
            item.get('agent_id'), item.get('invite_code'), item.get('display_name'), item.get('tags'),
            item.get('douyin_id'), item.get('fan_scale'), item.get('commission_rate'),
            item.get('total_invites'), item.get('total_paid_orders'), item.get('total_commission'), item.get('phone'),
        ])
    return buffer.getvalue()


def claim_referral_bonus(user_id: int, agent_id: int | None = None) -> dict[str, Any]:
    ensure_referral_p2_tables()
    settings = get_bonus_settings()
    bonus_beans = settings['bonus_beans']
    if bonus_beans <= 0:
        return {'claimed': False, 'message': '当前未开启达人专属福利'}
    with get_connection() as connection:
        exists = connection.execute('SELECT claim_id FROM referral_bonus_claims WHERE user_id = ?', [user_id]).fetchone()
        if exists:
            return {'claimed': False, 'message': '您已领取过达人专属福利'}
        binding = row_to_dict(connection.execute(
            'SELECT * FROM referral_bindings WHERE invitee_user_id = ?',
            [user_id]
        ).fetchone())
        if not binding:
            return {'claimed': False, 'message': '请先通过达人渠道进入后再领取'}
        resolved_agent_id = agent_id or binding.get('agent_id')
        connection.execute(
            'INSERT INTO referral_bonus_claims (user_id, agent_id, bonus_beans, bonus_days) VALUES (?, ?, ?, ?)',
            [user_id, resolved_agent_id, bonus_beans, settings['bonus_days']]
        )
        from bean_service import ensure_bean_tables, _get_account, _append_transaction
        ensure_bean_tables()
        account = _get_account(connection, user_id)
        new_balance = int(account.get('balance') or 0) + bonus_beans
        connection.execute(
            'UPDATE user_bean_accounts SET balance = ?, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?',
            [new_balance, user_id]
        )
        _append_transaction(
            connection, user_id, bonus_beans, new_balance, 'referral_bonus',
            f'达人渠道专属福利 +{bonus_beans} 星鼎豆', ''
        )
        connection.commit()
    return {
        'claimed': True,
        'bonus_beans': bonus_beans,
        'bonus_days': settings['bonus_days'],
        'message': f'已领取达人专属 {bonus_beans} 星鼎豆',
    }


def get_membership_compare() -> list[dict[str, Any]]:
    from bean_service import PLAN_BEAN_GRANT, PLAN_CATALOG, REPORT_BEAN_COST
    free_features = ['完整测评流程', '基础院校专业查询', '近2年分数线', '手动志愿模拟']
    return [
        {
            'plan_code': 'free',
            'plan_name': '免费体验',
            'price_text': '¥0',
            'duration_text': '长期',
            'bean_grant': 0,
            'report_cost': REPORT_BEAN_COST,
            'features': free_features,
        },
        {
            'plan_code': 'trial',
            'plan_name': PLAN_CATALOG['trial']['plan_name'],
            'price_text': '¥19.9',
            'duration_text': '30天',
            'bean_grant': PLAN_BEAN_GRANT['trial'],
            'report_cost': REPORT_BEAN_COST,
            'features': ['到账 2000 星鼎豆', '完整历年分数线', '深度测评报告', '智能推荐', 'AI 解读'],
        },
        {
            'plan_code': 'standard',
            'plan_name': PLAN_CATALOG['standard']['plan_name'],
            'price_text': '¥99',
            'duration_text': '365天',
            'bean_grant': PLAN_BEAN_GRANT['standard'],
            'report_cost': REPORT_BEAN_COST,
            'features': ['到账 12000 星鼎豆', '智能推荐不限次', '风险检测', 'AI 解读', 'PDF 导出'],
        },
        {
            'plan_code': 'premium',
            'plan_name': PLAN_CATALOG['premium']['plan_name'],
            'price_text': '¥168',
            'duration_text': '365天',
            'bean_grant': PLAN_BEAN_GRANT['premium'],
            'report_cost': REPORT_BEAN_COST,
            'features': ['到账 24000 星鼎豆', '金卡全部功能', '院校深度对比', '专属答疑通道'],
        },
    ]
