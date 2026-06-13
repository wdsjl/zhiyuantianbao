"""一分一段表：PDF/Excel 解析、入库与分数位次互查。"""

from __future__ import annotations

import io
import re
from typing import Any

from db import get_connection, row_to_dict, rows_to_dicts
SCORE_HEADER_ALIASES: tuple[str, ...] = ('分数', '文化课成绩', '高考成绩', '分值', '总分', '文化分')
SEGMENT_COUNT_ALIASES: tuple[str, ...] = ('本段人数', '人数', '同分人数', '本段人數', '段内人数')
CUMULATIVE_ALIASES: tuple[str, ...] = ('累计人数', '累计', '累计位次', '位次', '累计人數', '累计数')
SUBJECT_TYPE_ALIASES: tuple[str, ...] = ('物理', '历史', '理科', '文科', '物理类', '历史类')
BATCH_ALIASES: tuple[str, ...] = ('本科批', '专科批', '本科', '专科')


def _normalize_header(value: Any) -> str:
    return re.sub(r'\s+', '', str(value or '').strip())


def _match_alias(header: str, aliases: tuple[str, ...]) -> bool:
    normalized = _normalize_header(header)
    return any(alias in normalized or normalized in alias for alias in aliases)


def ensure_score_segment_tables() -> None:
    with get_connection() as connection:
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS score_rank_tables (
              table_id INTEGER PRIMARY KEY AUTOINCREMENT,
              province TEXT NOT NULL,
              year INTEGER NOT NULL,
              batch TEXT NOT NULL DEFAULT '',
              exam_type TEXT NOT NULL DEFAULT '普通类',
              subject_type TEXT NOT NULL DEFAULT '',
              title TEXT,
              source_file TEXT,
              row_count INTEGER NOT NULL DEFAULT 0,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              UNIQUE(province, year, batch, subject_type)
            )
            '''
        )
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS score_rank_segments (
              segment_id INTEGER PRIMARY KEY AUTOINCREMENT,
              table_id INTEGER NOT NULL,
              score INTEGER NOT NULL,
              segment_count INTEGER,
              cumulative_rank INTEGER NOT NULL,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              UNIQUE(table_id, score),
              FOREIGN KEY (table_id) REFERENCES score_rank_tables(table_id) ON DELETE CASCADE
            )
            '''
        )
        connection.execute(
            'CREATE INDEX IF NOT EXISTS idx_score_rank_lookup ON score_rank_segments(table_id, score)'
        )
        connection.execute(
            'CREATE INDEX IF NOT EXISTS idx_score_rank_tables_query ON score_rank_tables(province, year, batch)'
        )
        connection.commit()


def _to_int(value: Any) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).strip().replace(',', '')))
    except ValueError:
        return None


def _normalize_province(value: str) -> str:
    text = (value or '').strip()
    for suffix in ('省', '市', '自治区', '壮族自治区', '回族自治区', '维吾尔自治区'):
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def _province_variants(province: str) -> list[str]:
    base = _normalize_province(province)
    variants: list[str] = []
    for value in ((province or '').strip(), base, f'{base}省', f'{base}市'):
        if value and value not in variants:
            variants.append(value)
    return variants or [(province or '').strip()]


def _normalize_subject_type(value: str) -> str:
    text = (value or '').strip()
    if '物理' in text or text in ('理科', '理工'):
        return '物理'
    if '历史' in text or text in ('文科', '文史'):
        return '历史'
    return text


def detect_score_header_row_index(table: list[list[Any]]) -> int:
    for index, row in enumerate(table[:40]):
        mapping: dict[str, int] = {}
        for col_index, cell in enumerate(row or []):
            header = str(cell or '').strip()
            if not header:
                continue
            if _match_alias(header, SCORE_HEADER_ALIASES):
                mapping['score'] = col_index
            elif _match_alias(header, SEGMENT_COUNT_ALIASES):
                mapping['segment_count'] = col_index
            elif _match_alias(header, CUMULATIVE_ALIASES):
                mapping['cumulative_rank'] = col_index
        if 'score' in mapping and ('cumulative_rank' in mapping or 'segment_count' in mapping):
            return index
        if 'score' in mapping and len(mapping) >= 2:
            return index
    return -1


def _cell_value(row: list[Any] | tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _looks_like_score(value: int | None) -> bool:
    return value is not None and 100 <= value <= 900


def _looks_like_year(value: int | None) -> bool:
    return value is not None and 2020 <= value <= 2035


def parse_columnar_segments(
    table: list[list[Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """识别河南常见无表头五列：分数 | 本段人数 | 累计人数 | 年份 | 科类。"""
    rows: list[dict[str, Any]] = []
    meta: dict[str, Any] = {'year': None, 'subject_type': ''}
    matched = 0

    for row in table:
        if not row:
            continue
        first = _to_int(_cell_value(row, 0))
        second = _to_int(_cell_value(row, 1))
        third = _to_int(_cell_value(row, 2))
        fourth = _to_int(_cell_value(row, 3))
        fifth = str(_cell_value(row, 4) or '').strip()

        if _looks_like_year(first) and _normalize_subject_type(fifth):
            meta['year'] = first
            meta['subject_type'] = _normalize_subject_type(fifth)
            continue

        score = first
        if not _looks_like_score(score):
            continue

        segment_count: int | None = None
        cumulative_rank: int | None = None

        if third is not None and second is not None:
            if third >= second:
                segment_count = second
                cumulative_rank = third
            else:
                cumulative_rank = second
                segment_count = third
        elif second is not None:
            cumulative_rank = second

        if cumulative_rank is None:
            continue

        if _looks_like_year(fourth):
            meta['year'] = fourth
        if fifth:
            subject = _normalize_subject_type(fifth)
            if subject:
                meta['subject_type'] = subject

        rows.append({
            'score': int(score),
            'segment_count': segment_count,
            'cumulative_rank': int(cumulative_rank),
        })
        matched += 1

    return rows, meta


def parse_segments_from_matrix(
    table: list[list[Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    _, header_rows = parse_table_matrix(table)
    if header_rows:
        return header_rows, {}

    columnar_rows, meta = parse_columnar_segments(table)
    if columnar_rows:
        return columnar_rows, meta

    return [], {}


def validate_segment_rows(rows: list[dict[str, Any]]) -> None:
    if len(rows) < 10:
        raise ValueError(f'识别到的有效分数行过少（{len(rows)} 行），请检查文件列顺序是否为：分数、本段人数、累计人数')

    sorted_rows = sorted(rows, key=lambda item: item['score'], reverse=True)
    top = sorted_rows[0]
    bottom = sorted_rows[-1]
    if int(top['cumulative_rank']) > 20000:
        raise ValueError(
            f'最高分 {top["score"]} 的累计位次为 {top["cumulative_rank"]}，明显偏大。'
            '请确认第三列是「累计人数/位次」，第二列是「本段人数」。'
        )
    if int(bottom['cumulative_rank']) < int(top['cumulative_rank']):
        raise ValueError('累计位次未随分数递减而递增，请检查列顺序是否颠倒了')

    prev_cumulative = 0
    for row in sorted_rows:
        cumulative = int(row['cumulative_rank'])
        if cumulative < prev_cumulative:
            raise ValueError(f'分数 {row["score"]} 的累计位次小于更高分数，数据顺序异常')
        prev_cumulative = cumulative


def parse_table_matrix(
    table: list[list[Any]],
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    header_row_index = detect_score_header_row_index(table)
    if header_row_index < 0:
        return {}, []

    row = table[header_row_index]
    header_indexes: dict[str, int] = {}
    for col_index, cell in enumerate(row or []):
        header = str(cell or '').strip()
        if not header:
            continue
        if _match_alias(header, SCORE_HEADER_ALIASES):
            header_indexes['score'] = col_index
        elif _match_alias(header, SEGMENT_COUNT_ALIASES):
            header_indexes['segment_count'] = col_index
        elif _match_alias(header, CUMULATIVE_ALIASES):
            header_indexes['cumulative_rank'] = col_index

    rows: list[dict[str, Any]] = []
    for row in table[header_row_index + 1:]:
        if not row:
            continue

        def value(key: str) -> Any:
            idx = header_indexes.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        score = _to_int(value('score'))
        if score is None or score < 0 or score > 900:
            continue
        segment_count = _to_int(value('segment_count'))
        cumulative_rank = _to_int(value('cumulative_rank'))
        rows.append({
            'score': score,
            'segment_count': segment_count,
            'cumulative_rank': cumulative_rank,
        })
    return header_indexes, rows


def finalize_segment_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned = [row for row in rows if row.get('score') is not None]
    if not cleaned:
        return []

    has_cumulative = any(row.get('cumulative_rank') for row in cleaned)
    if has_cumulative:
        result = []
        for row in cleaned:
            cumulative = row.get('cumulative_rank')
            if cumulative is None:
                continue
            result.append({
                'score': int(row['score']),
                'segment_count': row.get('segment_count'),
                'cumulative_rank': int(cumulative),
            })
        return sorted(result, key=lambda item: item['score'], reverse=True)

    sorted_rows = sorted(cleaned, key=lambda item: item['score'], reverse=True)
    cumulative = 0
    result: list[dict[str, Any]] = []
    for row in sorted_rows:
        count = int(row.get('segment_count') or 0)
        cumulative += max(count, 0)
        result.append({
            'score': int(row['score']),
            'segment_count': count or None,
            'cumulative_rank': cumulative,
        })
    return result


def extract_segments_from_pdf(content: bytes) -> list[dict[str, Any]]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError('服务器未安装 pdfplumber，请执行 pip install pdfplumber') from exc

    collected: list[dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                rows, _ = parse_segments_from_matrix(table)
                collected.extend(rows)
            if collected:
                continue
            text = page.extract_text() or ''
            collected.extend(_extract_segments_from_text(text))
    return finalize_segment_rows(collected)


def _extract_segments_from_text(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    patterns = [
        re.compile(r'(\d{2,3})\s+(\d{1,5})\s+(\d{1,6})'),
        re.compile(r'(\d{2,3})\s+(\d{1,6})'),
    ]
    for line in text.splitlines():
        line = line.strip()
        if not line or not re.search(r'\d', line):
            continue
        match3 = patterns[0].search(line)
        if match3:
            score, segment_count, cumulative = match3.groups()
            rows.append({
                'score': int(score),
                'segment_count': int(segment_count),
                'cumulative_rank': int(cumulative),
            })
            continue
        match2 = patterns[1].search(line)
        if match2:
            score, cumulative = match2.groups()
            rows.append({
                'score': int(score),
                'segment_count': None,
                'cumulative_rank': int(cumulative),
            })
    return rows


def extract_segments_from_spreadsheet(content: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    matrix: list[list[Any]] = []
    lower = (filename or '').lower()
    if lower.endswith('.csv'):
        import csv
        text = content.decode('utf-8-sig', errors='replace')
        matrix = list(csv.reader(io.StringIO(text)))
    else:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
    parsed, meta = parse_segments_from_matrix(matrix)
    finalized = finalize_segment_rows(parsed)
    if finalized:
        validate_segment_rows(finalized)
    return finalized, meta


def extract_segments_from_file(content: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    lower = (filename or '').lower()
    if lower.endswith('.pdf') or content[:4] == b'%PDF':
        segments = extract_segments_from_pdf(content)
        if segments:
            validate_segment_rows(segments)
        return segments, {}
    if lower.endswith(('.xlsx', '.xls', '.csv')):
        return extract_segments_from_spreadsheet(content, filename)
    raise ValueError('仅支持 PDF、Excel 或 CSV 格式的一分一段表')


def upsert_score_rank_table(
    *,
    province: str,
    year: int,
    batch: str = '',
    exam_type: str = '普通类',
    subject_type: str = '',
    title: str = '',
    source_file: str = '',
    segments: list[dict[str, Any]],
) -> dict[str, Any]:
    ensure_score_segment_tables()
    province = _normalize_province(province)
    subject_type = _normalize_subject_type(subject_type)
    if not segments:
        raise ValueError('未能从文件中识别出一分一段数据，请检查表头是否含「分数」「累计人数/位次」等列')

    with get_connection() as connection:
        existing = row_to_dict(
            connection.execute(
                '''
                SELECT table_id FROM score_rank_tables
                WHERE province = ? AND year = ? AND batch = ? AND subject_type = ?
                ''',
                [province, year, batch or '', subject_type or ''],
            ).fetchone()
        )
        if existing:
            table_id = int(existing['table_id'])
            connection.execute(
                '''
                UPDATE score_rank_tables
                SET title = ?, source_file = ?, row_count = ?, exam_type = ?, updated_at = CURRENT_TIMESTAMP
                WHERE table_id = ?
                ''',
                [title or source_file, source_file, len(segments), exam_type, table_id],
            )
            connection.execute('DELETE FROM score_rank_segments WHERE table_id = ?', [table_id])
        else:
            cursor = connection.execute(
                '''
                INSERT INTO score_rank_tables (
                  province, year, batch, exam_type, subject_type, title, source_file, row_count
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                [province, year, batch or '', exam_type, subject_type or '', title or source_file, source_file, len(segments)],
            )
            table_id = int(cursor.lastrowid)

        for item in segments:
            connection.execute(
                '''
                INSERT INTO score_rank_segments (table_id, score, segment_count, cumulative_rank)
                VALUES (?, ?, ?, ?)
                ''',
                [table_id, item['score'], item.get('segment_count'), item['cumulative_rank']],
            )
        connection.commit()
    return {
        'table_id': table_id,
        'province': province,
        'year': year,
        'batch': batch,
        'subject_type': subject_type or '',
        'row_count': len(segments),
        'score_min': min(item['score'] for item in segments),
        'score_max': max(item['score'] for item in segments),
    }


def import_score_segment_file(
    content: bytes,
    filename: str,
    *,
    province: str,
    year: int,
    batch: str = '',
    exam_type: str = '普通类',
    subject_type: str = '',
    title: str = '',
) -> dict[str, Any]:
    province = _normalize_province(province)
    subject_type = _normalize_subject_type(subject_type)
    if not subject_type:
        lower_name = (filename or '').lower()
        if '物理' in lower_name or '理科' in lower_name:
            subject_type = '物理'
        elif '历史' in lower_name or '文科' in lower_name:
            subject_type = '历史'
    segments, meta = extract_segments_from_file(content, filename)
    file_year = meta.get('year')
    if file_year and not year:
        year = int(file_year)
    file_subject = _normalize_subject_type(str(meta.get('subject_type') or ''))
    if file_subject and not subject_type:
        subject_type = file_subject
    return upsert_score_rank_table(
        province=province,
        year=year,
        batch=batch or '',
        exam_type=exam_type,
        subject_type=subject_type,
        title=title,
        source_file=filename,
        segments=segments,
    )


def _query_score_rank_table(
    province_variants: list[str],
    *,
    year: int | None,
    batch: str,
    subject_type: str,
) -> dict[str, Any] | None:
    placeholders = ','.join(['?'] * len(province_variants))
    sql = f'SELECT * FROM score_rank_tables WHERE province IN ({placeholders})'
    params: list[Any] = [*province_variants]
    if year:
        sql += ' AND year = ?'
        params.append(year)
    if batch:
        sql += ' AND batch = ?'
        params.append(batch)
    if subject_type:
        sql += ' AND subject_type = ?'
        params.append(subject_type)
    sql += ' ORDER BY year DESC, table_id DESC LIMIT 1'
    with get_connection() as connection:
        return row_to_dict(connection.execute(sql, params).fetchone())


def find_score_rank_table(
    province: str,
    year: int | None = None,
    batch: str = '',
    subject_type: str = '',
) -> dict[str, Any] | None:
    ensure_score_segment_tables()
    variants = _province_variants(province)
    normalized_subject = _normalize_subject_type(subject_type)
    normalized_batch = (batch or '').strip()

    batch_candidates: list[str] = []
    for value in (normalized_batch, ''):
        if value not in batch_candidates:
            batch_candidates.append(value)
    if normalized_batch in ('本科批', '本科'):
        for value in ('', '本科批', '本科'):
            if value not in batch_candidates:
                batch_candidates.append(value)

    subject_candidates: list[str] = []
    for value in (normalized_subject, ''):
        if value not in subject_candidates:
            subject_candidates.append(value)

    for batch_value in batch_candidates:
        for subject_value in subject_candidates:
            table = _query_score_rank_table(
                variants,
                year=year,
                batch=batch_value,
                subject_type=subject_value,
            )
            if table:
                return table
    return None


def lookup_rank_by_score(
    province: str,
    score: int,
    *,
    year: int | None = None,
    batch: str = '',
    subject_type: str = '',
) -> int | None:
    table = find_score_rank_table(province, year, batch, subject_type)
    if not table:
        return None
    with get_connection() as connection:
        exact = row_to_dict(
            connection.execute(
                '''
                SELECT cumulative_rank FROM score_rank_segments
                WHERE table_id = ? AND score = ?
                ''',
                [table['table_id'], int(score)],
            ).fetchone()
        )
        if exact:
            return int(exact['cumulative_rank'])
        lower = row_to_dict(
            connection.execute(
                '''
                SELECT score, cumulative_rank FROM score_rank_segments
                WHERE table_id = ? AND score <= ?
                ORDER BY score DESC
                LIMIT 1
                ''',
                [table['table_id'], int(score)],
            ).fetchone()
        )
        if lower:
            return int(lower['cumulative_rank'])
        higher = row_to_dict(
            connection.execute(
                '''
                SELECT score, cumulative_rank FROM score_rank_segments
                WHERE table_id = ? AND score >= ?
                ORDER BY score ASC
                LIMIT 1
                ''',
                [table['table_id'], int(score)],
            ).fetchone()
        )
        if higher:
            return int(higher['cumulative_rank'])
    return None


def lookup_score_by_rank(
    province: str,
    rank: int,
    *,
    year: int | None = None,
    batch: str = '',
    subject_type: str = '',
) -> int | None:
    table = find_score_rank_table(province, year, batch, subject_type)
    if not table:
        return None
    with get_connection() as connection:
        row = row_to_dict(
            connection.execute(
                '''
                SELECT score, cumulative_rank FROM score_rank_segments
                WHERE table_id = ? AND cumulative_rank >= ?
                ORDER BY cumulative_rank ASC, score DESC
                LIMIT 1
                ''',
                [table['table_id'], int(rank)],
            ).fetchone()
        )
        if row:
            return int(row['score'])
    return None


def infer_subject_type_from_combination(subject_combination: str) -> str:
    combo = (subject_combination or '').replace('/', '+').replace('、', '+')
    if '物理' in combo or combo.startswith('物') or '+物' in combo or '物化' in combo:
        return '物理'
    if '历史' in combo or combo.startswith('历') or '+历' in combo or '史政' in combo:
        return '历史'
    return ''


def list_score_rank_tables(limit: int = 50) -> list[dict[str, Any]]:
    ensure_score_segment_tables()
    with get_connection() as connection:
        return rows_to_dicts(
            connection.execute(
                'SELECT * FROM score_rank_tables ORDER BY year DESC, province ASC, table_id DESC LIMIT ?',
                [int(limit)],
            ).fetchall()
        )
